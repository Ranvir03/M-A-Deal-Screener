"""
M&A Deal Screener
=================

A professional-grade M&A screening tool built with Streamlit.

Design goals:
  * Works for PUBLIC and PRIVATE targets (small-firm reality: most targets are private).
  * Three data sources, all producing the SAME company "snapshot":
        1. Manual entry   -> type the numbers straight from a 10-K / CIM / tax return,
                             with an add-backs section that builds Adjusted EBITDA.
        2. SEC EDGAR      -> free, audited filing data pulled from XBRL (US public co's).
        3. Yahoo Finance  -> quick public-market pull for a fast first look.
  * A defensible model: real unlevered free cash flow, correct enterprise->equity
    bridge, and foregone interest on cash in the accretion math.

Nothing here is a substitute for full diligence or a banker's judgment on
normalization / add-backs. See the "Assumptions & Limitations" panel in the app.
"""

import io
import json
from datetime import date, datetime

import requests
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import streamlit as st

# SEC requires a descriptive User-Agent with contact info on every request.
SEC_HEADERS = {"User-Agent": "MA-Deal-Screener contact@example.com"}
SEC_TIMEOUT = 30


# ======================================================================
# Snapshot schema
# ----------------------------------------------------------------------
# Every data source returns a dict with these keys. Downstream model code
# only ever touches this dict, so sources are fully interchangeable.
# ======================================================================

SNAPSHOT_FIELDS = [
    "ticker", "name", "sector", "industry",
    "price", "shares_out",
    "revenue", "ebitda", "add_backs",
    "da", "capex", "nwc_change",
    "net_income", "interest_expense", "tax_rate_reported",
    "total_debt", "cash",
    "beta", "revenue_growth",
    "source",
]


def finalize_snapshot(raw: dict) -> dict:
    """Fill derived fields and validate. Accepts a partial raw dict."""
    s = {k: raw.get(k) for k in SNAPSHOT_FIELDS}

    # sensible fallbacks for optional descriptive fields
    s["name"] = s["name"] or s.get("ticker") or "Company"
    s["sector"] = s["sector"] or "N/A"
    s["industry"] = s["industry"] or "N/A"

    def num(x, default=None):
        try:
            return float(x) if x is not None and x != "" else default
        except (TypeError, ValueError):
            return default

    for k in ["price", "shares_out", "revenue", "ebitda", "add_backs", "da",
              "capex", "nwc_change", "net_income", "interest_expense",
              "total_debt", "cash", "beta", "revenue_growth", "tax_rate_reported"]:
        s[k] = num(s[k])

    s["add_backs"] = s["add_backs"] or 0.0
    s["total_debt"] = s["total_debt"] or 0.0
    s["cash"] = s["cash"] or 0.0

    # Adjusted EBITDA = reported EBITDA + add-backs (the number deals price on)
    s["adjusted_ebitda"] = (s["ebitda"] or 0.0) + s["add_backs"]

    # EBIT = EBITDA - D&A when we have D&A; otherwise fall back to EBITDA
    if s["ebitda"] is not None and s["da"] is not None:
        s["ebit"] = s["ebitda"] - s["da"]
    else:
        s["ebit"] = s["ebitda"]

    s["net_debt"] = s["total_debt"] - s["cash"]

    if s["price"] and s["shares_out"]:
        s["market_cap"] = s["price"] * s["shares_out"]
        s["enterprise_value"] = s["market_cap"] + s["net_debt"]
    else:
        s["market_cap"] = None
        s["enterprise_value"] = None

    ev = s["enterprise_value"]
    s["ev_rev"] = ev / s["revenue"] if ev and s["revenue"] else None
    s["ev_ebitda"] = ev / s["adjusted_ebitda"] if ev and s["adjusted_ebitda"] else None

    return s


def snapshot_completeness(s: dict) -> list:
    """Return a list of human-readable warnings about missing critical inputs."""
    warns = []
    if not s.get("revenue"):
        warns.append("Revenue is missing.")
    if not s.get("ebitda"):
        warns.append("EBITDA is missing.")
    if not s.get("shares_out"):
        warns.append("Shares outstanding is missing (needed for per-share / EPS math).")
    if not s.get("price"):
        warns.append("Share price is missing (needed for market cap, premium, accretion).")
    if s.get("da") is None:
        warns.append("D&A not provided — DCF will assume capex ≈ D&A and use a proxy.")
    return warns


# ======================================================================
# Data source: SEC EDGAR (free, audited XBRL filing data)
# ======================================================================

@st.cache_data(ttl=3600, show_spinner=False)
def _edgar_ticker_map() -> dict:
    """Map upper-case ticker -> zero-padded 10-digit CIK."""
    r = requests.get("https://www.sec.gov/files/company_tickers.json",
                     headers=SEC_HEADERS, timeout=SEC_TIMEOUT)
    r.raise_for_status()
    out = {}
    for row in r.json().values():
        out[row["ticker"].upper()] = str(row["cik_str"]).zfill(10)
    return out


@st.cache_data(ttl=3600, show_spinner=False)
def _edgar_company_facts(cik: str) -> dict:
    r = requests.get(f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json",
                     headers=SEC_HEADERS, timeout=SEC_TIMEOUT)
    r.raise_for_status()
    return r.json()


def _parse_date(s: str) -> date:
    y, m, d = map(int, s.split("-"))
    return date(y, m, d)


def _latest_annual_flow(unit_entries: list):
    """Most recent ~full-year (350-380 day) value from an income/cash-flow concept."""
    cand = []
    for x in unit_entries:
        if not x.get("start"):
            continue
        span = (_parse_date(x["end"]) - _parse_date(x["start"])).days
        if 350 <= span <= 380:
            cand.append(x)
    if not cand:
        return None
    cand.sort(key=lambda x: x["end"])
    return cand[-1]["val"]


def _latest_instant(unit_entries: list):
    """Most recent point-in-time value from a balance-sheet concept."""
    if not unit_entries:
        return None
    entries = sorted(unit_entries, key=lambda x: x["end"])
    return entries[-1]["val"]


def _pick_concept(gaap: dict, names: list, mode: str):
    """Try each candidate XBRL tag; return the value from whichever is most recent."""
    best_val, best_end = None, None
    for name in names:
        node = gaap.get(name)
        if not node:
            continue
        units = node.get("units", {})
        series = units.get("USD") or units.get("shares")
        if not series:
            continue
        if mode == "flow":
            entries = [x for x in series if x.get("start")
                       and 350 <= (_parse_date(x["end"]) - _parse_date(x["start"])).days <= 380]
        else:
            entries = series
        if not entries:
            continue
        latest = max(entries, key=lambda x: x["end"])
        if best_end is None or latest["end"] > best_end:
            best_end, best_val = latest["end"], latest["val"]
    return best_val


def fetch_from_edgar(ticker: str, price: float = None) -> dict:
    """Pull fundamentals from the latest 10-K via EDGAR XBRL. Price is a market
    input the user supplies (filings don't contain a share price)."""
    ticker = ticker.upper().strip()
    cik = _edgar_ticker_map().get(ticker)
    if not cik:
        raise ValueError(f"'{ticker}' not found in SEC EDGAR (US public filers only).")

    facts = _edgar_company_facts(cik)
    gaap = facts["facts"].get("us-gaap", {})
    dei = facts["facts"].get("dei", {})
    name = facts.get("entityName", ticker)

    revenue = _pick_concept(gaap, [
        "RevenueFromContractWithCustomerExcludingAssessedTax",
        "RevenueFromContractWithCustomerIncludingAssessedTax",
        "Revenues", "SalesRevenueNet",
    ], "flow")
    net_income = _pick_concept(gaap, ["NetIncomeLoss", "ProfitLoss"], "flow")
    operating_income = _pick_concept(gaap, ["OperatingIncomeLoss"], "flow")
    da = _pick_concept(gaap, [
        "DepreciationDepletionAndAmortization",
        "DepreciationAmortizationAndAccretionNet",
        "DepreciationAndAmortization", "Depreciation",
    ], "flow")
    capex = _pick_concept(gaap, [
        "PaymentsToAcquirePropertyPlantAndEquipment",
        "PaymentsToAcquireProductiveAssets",
    ], "flow")
    interest = _pick_concept(gaap, [
        "InterestExpense", "InterestExpenseDebt", "InterestAndDebtExpense",
    ], "flow")
    pretax = _pick_concept(gaap, [
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments",
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest",
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesDomestic",
    ], "flow")
    tax_exp = _pick_concept(gaap, ["IncomeTaxExpenseBenefit"], "flow")

    cash = _pick_concept(gaap, [
        "CashAndCashEquivalentsAtCarryingValue",
        "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents",
    ], "instant")
    lt_debt = _pick_concept(gaap, ["LongTermDebtNoncurrent", "LongTermDebt"], "instant") or 0
    cur_debt = _pick_concept(gaap, ["LongTermDebtCurrent", "DebtCurrent"], "instant") or 0
    total_debt = (lt_debt or 0) + (cur_debt or 0)

    shares = None
    node = dei.get("EntityCommonStockSharesOutstanding")
    if node:
        shares = _latest_instant(node.get("units", {}).get("shares", []))
    if not shares:
        shares = _pick_concept(gaap, ["CommonStockSharesOutstanding"], "instant")

    # EBITDA is rarely tagged directly -> build it, with fallbacks because many
    # companies (e.g. energy/industrials) don't tag OperatingIncomeLoss.
    ebitda = None
    if operating_income is not None and da is not None:
        ebitda = operating_income + da                                  # EBIT + D&A
    elif pretax is not None and da is not None:
        ebitda = pretax + (interest or 0) + da                          # pre-tax + interest + D&A
    elif net_income is not None and da is not None:
        ebitda = net_income + (tax_exp or 0) + (interest or 0) + da     # bottom-up
    elif operating_income is not None:
        ebitda = operating_income

    return finalize_snapshot({
        "ticker": ticker, "name": name,
        "price": price, "shares_out": shares,
        "revenue": revenue, "ebitda": ebitda, "add_backs": 0.0,
        "da": da, "capex": capex, "nwc_change": None,
        "net_income": net_income, "interest_expense": interest,
        "total_debt": total_debt, "cash": cash,
        "beta": None, "revenue_growth": None,
        "source": "SEC EDGAR (10-K XBRL)",
    })


# ======================================================================
# Data source: Yahoo Finance (quick public pull)
# ======================================================================

def fetch_from_yahoo(ticker: str) -> dict:
    import yfinance as yf
    info = yf.Ticker(ticker).info or {}

    def g(*keys):
        for k in keys:
            v = info.get(k)
            if v is not None:
                return v
        return None

    return finalize_snapshot({
        "ticker": ticker,
        "name": g("longName", "shortName") or ticker,
        "sector": g("sector"), "industry": g("industry"),
        "price": g("currentPrice", "regularMarketPrice"),
        "shares_out": g("sharesOutstanding"),
        "revenue": g("totalRevenue"),
        "ebitda": g("ebitda"),
        "add_backs": 0.0,
        "da": None, "capex": None, "nwc_change": None,
        "net_income": g("netIncomeToCommon"),
        "interest_expense": g("interestExpense"),
        "total_debt": g("totalDebt") or 0,
        "cash": g("totalCash") or 0,
        "beta": g("beta"),
        "revenue_growth": g("revenueGrowth"),
        "source": "Yahoo Finance",
    })


# ======================================================================
# Valuation building blocks
# ======================================================================

def estimate_cost_of_debt(c: dict, fallback: float = 0.055) -> float:
    interest, debt = c.get("interest_expense"), c.get("total_debt")
    if interest and debt and debt > 0:
        r = abs(interest) / debt
        if 0.0 < r < 0.20:
            return r
    return fallback


def estimate_cost_of_equity(c: dict, rf: float, erp: float) -> float:
    beta = c.get("beta") or 1.0
    return rf + beta * erp


def estimate_wacc(c: dict, tax_rate: float, rf: float, erp: float) -> float:
    ke = estimate_cost_of_equity(c, rf, erp)
    kd = estimate_cost_of_debt(c)
    debt = c.get("total_debt") or 0
    equity = c.get("market_cap") or 0
    total = debt + equity
    if total <= 0:            # private target with no market cap -> use cost of equity
        return ke
    return (equity / total) * ke + (debt / total) * kd * (1 - tax_rate)


def dcf_valuation(t: dict, tax_rate: float, wacc: float,
                  years: int = 5, terminal_growth: float = 0.025,
                  projection: list = None) -> dict:
    """Proper unlevered-FCF DCF with an enterprise->equity bridge.

    UFCF = EBIT*(1-tax) + D&A - Capex - change in net working capital.
    Discounting UFCF at WACC gives ENTERPRISE value; subtract net debt for EQUITY value.

    If `projection` is supplied (a list of {"revenue","ebitda"} dicts, one per
    forecast year), those explicit figures drive the forecast instead of a single
    flat growth rate — this is the multi-year path.
    """
    empty = {"enterprise_value": None, "equity_value": None, "fcf_list": None,
             "wacc": wacc, "terminal_value": None, "assumed_da": False, "revenue_proj": None}
    revenue = t.get("revenue")
    ebitda = t.get("adjusted_ebitda") or t.get("ebitda")
    if not revenue or not ebitda or not wacc or wacc <= terminal_growth:
        return empty

    assumed_da = t.get("da") is None
    da = t.get("da") if t.get("da") is not None else revenue * 0.04      # proxy if unknown
    capex = t.get("capex") if t.get("capex") is not None else da         # maint. capex ≈ D&A
    nwc = t.get("nwc_change") or 0.0
    growth = t.get("revenue_growth") if t.get("revenue_growth") is not None else 0.05

    # base-year ratios (held constant when scaling forward)
    da_pct, capex_pct, nwc_pct = da / revenue, capex / revenue, nwc / revenue
    ebit_margin = (ebitda - da) / revenue

    fcf_list, rev_proj = [], []
    if projection:                                    # explicit multi-year forecast
        for yr in projection:
            rev_y = yr["revenue"]
            ebitda_y = yr.get("ebitda") or rev_y * (ebitda / revenue)
            da_y, capex_y, nwc_y = rev_y * da_pct, rev_y * capex_pct, rev_y * nwc_pct
            ufcf = (ebitda_y - da_y) * (1 - tax_rate) + da_y - capex_y - nwc_y
            fcf_list.append(ufcf)
            rev_proj.append(rev_y)
        years = len(projection)
    else:                                             # single flat-growth forecast
        rev = revenue
        for _ in range(years):
            rev *= (1 + growth)
            ufcf = (rev * ebit_margin) * (1 - tax_rate) + rev * da_pct - rev * capex_pct - rev * nwc_pct
            fcf_list.append(ufcf)
            rev_proj.append(rev)

    terminal_value = fcf_list[-1] * (1 + terminal_growth) / (wacc - terminal_growth)
    ev = sum(f / ((1 + wacc) ** i) for i, f in enumerate(fcf_list, start=1))
    ev += terminal_value / ((1 + wacc) ** years)
    equity_value = ev - (t.get("net_debt") or 0)

    return {"enterprise_value": ev, "equity_value": equity_value, "fcf_list": fcf_list,
            "wacc": wacc, "terminal_value": terminal_value, "assumed_da": assumed_da,
            "revenue_proj": rev_proj}


def build_offer(t: dict, mode: str, premium=None, ev_multiple=None, ev_abs=None) -> dict:
    """Offer can be set three ways so PRIVATE targets work, not just public ones:
       - 'premium'  : premium to current share price (public target)
       - 'multiple' : purchase EV = multiple x Adjusted EBITDA
       - 'ev_abs'   : purchase EV entered directly ($)
    """
    net_debt = t.get("net_debt") or 0
    price, shares = t.get("price"), t.get("shares_out")
    eq_now = price * shares if price and shares else None

    if mode == "premium":
        if not price or not shares:
            raise ValueError("Premium-to-market offer needs the target's share price and share count.")
        offer_price = price * (1 + premium)
        eq_offer = offer_price * shares
        ev_offer = eq_offer + net_debt
    elif mode == "multiple":
        adj = t.get("adjusted_ebitda")
        if not adj:
            raise ValueError("EV/EBITDA offer needs the target's (Adjusted) EBITDA.")
        ev_offer = ev_multiple * adj
        eq_offer = ev_offer - net_debt
        offer_price = eq_offer / shares if shares else None
    else:  # ev_abs
        ev_offer = ev_abs
        eq_offer = ev_offer - net_debt
        offer_price = eq_offer / shares if shares else None

    implied_premium = (eq_offer / eq_now - 1) if eq_now else None
    premium_amount = (eq_offer - eq_now) if eq_now else None
    implied_ev_ebitda = ev_offer / t["adjusted_ebitda"] if t.get("adjusted_ebitda") else None

    return {
        "current_price": price, "offer_price": offer_price,
        "equity_value_current": eq_now, "equity_value_offer": eq_offer,
        "ev_offer": ev_offer, "premium_pct": implied_premium,
        "premium_amount": premium_amount, "implied_ev_ebitda": implied_ev_ebitda,
    }


def build_synergies(t: dict, cost_syn_pct: float, rev_syn_pct: float,
                    syn_margin: float, tax_rate: float) -> dict:
    revenue = t.get("revenue") or 0
    cost_syn = cost_syn_pct * revenue
    rev_syn_profit = (rev_syn_pct * revenue) * syn_margin
    ebit = cost_syn + rev_syn_profit
    return {"annual_cost_synergies": cost_syn, "annual_rev_synergies": rev_syn_pct * revenue,
            "annual_synergy_ebit": ebit, "annual_synergy_after_tax": ebit * (1 - tax_rate)}


def synergy_npv(annual_after_tax: float, wacc: float, horizon: int = 10,
                ramp=(0.25, 0.60, 1.0), cost_to_achieve: float = 0.0) -> dict:
    """NPV of after-tax synergies over a horizon, phased in by `ramp`, net of a
    one-time (year-1) cost-to-achieve."""
    if not annual_after_tax or annual_after_tax <= 0 or not wacc or wacc <= 0:
        return {"synergy_npv": None}
    ramp = list(ramp) + [ramp[-1]] * max(0, horizon - len(ramp))
    npv = sum(annual_after_tax * ramp[i - 1] / ((1 + wacc) ** i) for i in range(1, horizon + 1))
    npv -= (cost_to_achieve or 0.0) / (1 + wacc)      # cost-to-achieve hits year 1
    return {"synergy_npv": npv}


def build_pro_forma(buyer, target, offer, syn, tax_rate,
                    debt_pct, stock_pct, cash_yield,
                    synergy_ramp=(0.25, 0.60, 1.0), transaction_fee=0.0, cost_to_achieve=0.0,
                    intangible_pct=0.0, intangible_life=15,
                    financing_fee_pct=0.0, financing_life=7) -> dict:
    deal_value = offer["equity_value_offer"]
    new_debt = deal_value * debt_pct
    equity_issued = deal_value * stock_pct
    financing_fee = new_debt * (financing_fee_pct or 0.0)

    # Sources & Uses: fees and cost-to-achieve are funded with cash.
    total_uses = deal_value + (transaction_fee or 0.0) + (cost_to_achieve or 0.0) + financing_fee
    cash_used = max(total_uses - new_debt - equity_issued, 0.0)

    pf_net_debt = (buyer.get("net_debt") or 0) + (target.get("net_debt") or 0) + new_debt
    pf_ebitda = ((buyer.get("adjusted_ebitda") or 0) + (target.get("adjusted_ebitda") or 0)
                 + syn["annual_synergy_ebit"])

    cod = estimate_cost_of_debt(buyer)
    new_debt_interest_at = new_debt * cod * (1 - tax_rate)
    foregone_at = cash_used * cash_yield * (1 - tax_rate)     # foregone interest on cash spent

    # Purchase accounting: new intangibles amortized over life (after-tax EPS drag)
    intangibles = deal_value * (intangible_pct or 0.0)
    intangible_amort_at = (intangibles / intangible_life) * (1 - tax_rate) if intangible_life else 0.0
    # Financing fees amortized over the debt life (after-tax EPS drag)
    fin_fee_amort_at = (financing_fee / financing_life) * (1 - tax_rate) if financing_life else 0.0

    buyer_ni = buyer.get("net_income") or 0
    target_ni = target.get("net_income") or 0
    synergy_run_at = syn["annual_synergy_after_tax"]
    y1_factor = synergy_ramp[0] if synergy_ramp else 1.0
    synergy_y1_at = synergy_run_at * y1_factor

    recurring_drag = new_debt_interest_at + foregone_at + intangible_amort_at + fin_fee_amort_at
    pf_ni_y1 = buyer_ni + target_ni + synergy_y1_at - recurring_drag        # phased (headline)
    pf_ni_rr = buyer_ni + target_ni + synergy_run_at - recurring_drag       # full run-rate

    buyer_shares = buyer.get("shares_out") or 0
    buyer_price = buyer.get("price")
    new_shares = (equity_issued / buyer_price) if (stock_pct > 0 and buyer_price) else 0.0
    pf_shares = buyer_shares + new_shares

    buyer_eps = (buyer_ni / buyer_shares) if (buyer_ni and buyer_shares > 0) else None
    pf_eps_y1 = (pf_ni_y1 / pf_shares) if pf_shares > 0 else None
    pf_eps_rr = (pf_ni_rr / pf_shares) if pf_shares > 0 else None

    def acc(eps):
        return ((eps - buyer_eps) / buyer_eps) if (buyer_eps not in (None, 0) and eps is not None) else None

    leverage = (pf_net_debt / pf_ebitda) if pf_ebitda and pf_ebitda > 0 else None

    return {"deal_value": deal_value, "new_debt": new_debt, "equity_issued_value": equity_issued,
            "cash_used": cash_used, "new_shares_issued": new_shares, "total_uses": total_uses,
            "transaction_fee": transaction_fee or 0.0, "cost_to_achieve": cost_to_achieve or 0.0,
            "financing_fee": financing_fee, "intangibles": intangibles,
            "intangible_amort_at": intangible_amort_at, "fin_fee_amort_at": fin_fee_amort_at,
            "pf_net_debt": pf_net_debt, "pf_ebitda": pf_ebitda, "pf_leverage": leverage,
            "buyer_eps": buyer_eps, "pf_eps": pf_eps_y1, "pf_eps_runrate": pf_eps_rr,
            "accretion_pct": acc(pf_eps_y1), "accretion_pct_runrate": acc(pf_eps_rr),
            "new_debt_interest_at": new_debt_interest_at, "foregone_interest_at": foregone_at,
            "synergy_after_tax": synergy_run_at, "synergy_after_tax_y1": synergy_y1_at}


def comps_valuation(target: dict, peers: list) -> dict:
    ev_ebitda = [p["ev_ebitda"] for p in peers if p.get("ev_ebitda") and p["ev_ebitda"] > 0]
    ev_rev = [p["ev_rev"] for p in peers if p.get("ev_rev") and p["ev_rev"] > 0]
    med_ee = float(np.median(ev_ebitda)) if ev_ebitda else None
    med_er = float(np.median(ev_rev)) if ev_rev else None
    nd = target.get("net_debt") or 0
    ev_from_ebitda = med_ee * target["adjusted_ebitda"] if med_ee and target.get("adjusted_ebitda") else None
    ev_from_rev = med_er * target["revenue"] if med_er and target.get("revenue") else None
    return {"median_ev_ebitda": med_ee, "median_ev_rev": med_er,
            "implied_ev_from_ebitda": ev_from_ebitda,
            "implied_ev_from_rev": ev_from_rev,
            "implied_equity_from_ebitda": (ev_from_ebitda - nd) if ev_from_ebitda else None,
            "implied_equity_from_rev": (ev_from_rev - nd) if ev_from_rev else None,
            "n_peers": len(peers)}


# ======================================================================
# Screening flags & verdict  (illustrative — NOT a recommendation)
# ======================================================================

def screening_flags(buyer, target) -> dict:
    sector_match = buyer.get("sector") == target.get("sector") and buyer.get("sector") not in (None, "N/A")
    industry_match = buyer.get("industry") == target.get("industry") and buyer.get("industry") not in (None, "N/A")
    strat = (4 if sector_match else 0) + (4 if industry_match else 0)
    reg = (3 if (buyer.get("market_cap") or 0) > 200e9 else 0) + (2 if sector_match else 0)
    return {"strategic_score": strat, "sector_match": sector_match,
            "industry_match": industry_match, "reg_score": reg, "reg_high_risk": reg >= 4}


def build_verdict(offer, syn, syn_npv, pf, dcf) -> dict:
    acc, lev, premium = pf["accretion_pct"], pf["pf_leverage"], offer["premium_pct"]
    prem_amt = offer.get("premium_amount")
    syn_at, syn_npv_val = syn["annual_synergy_after_tax"], syn_npv["synergy_npv"]

    cov_npv = (syn_npv_val / prem_amt) if (prem_amt and prem_amt > 0 and syn_npv_val) else None
    dcf_equity = dcf.get("equity_value")
    dcf_vs_offer = (dcf_equity - offer["equity_value_offer"]) if dcf_equity is not None else None

    premium_ok = (premium is None) or (premium <= 0.35)
    premium_soft = (premium is None) or (premium <= 0.45)
    if acc is None and lev is None:
        verdict = "Inconclusive – key public-market inputs missing (private buyer?)"
    elif (acc is not None and acc > 0.05) and (lev is None or lev <= 3.5) and premium_ok:
        verdict = "Attractive screening profile"
    elif (acc is None or acc > 0) and (lev is None or lev <= 4.5) and premium_soft:
        verdict = "Mixed – hinges on strategic rationale & synergy delivery"
    else:
        verdict = "Financially challenging on these assumptions"

    return {"accretion_pct": acc, "pf_leverage": lev, "premium_pct": premium,
            "synergy_coverage_npv": cov_npv, "synergy_npv": syn_npv_val,
            "dcf_equity": dcf_equity, "dcf_vs_offer": dcf_vs_offer, "verdict": verdict}


# ======================================================================
# Scenario engine
# ======================================================================

def run_case(buyer, target, offer_kwargs, debt_pct, stock_pct, tax_rate,
             cost_syn, rev_syn, syn_margin, cash_yield, rf, erp, term_growth,
             wacc_override=None, synergy_ramp=(0.25, 0.60, 1.0),
             transaction_fee=0.0, cost_to_achieve=0.0,
             intangible_pct=0.0, intangible_life=15,
             financing_fee_pct=0.0, financing_life=7, projection=None):
    offer = build_offer(target, **offer_kwargs)
    syn = build_synergies(target, cost_syn, rev_syn, syn_margin, tax_rate)
    wacc = wacc_override if wacc_override else estimate_wacc(target, tax_rate, rf, erp)
    sn = synergy_npv(syn["annual_synergy_after_tax"], wacc,
                     ramp=synergy_ramp, cost_to_achieve=cost_to_achieve)
    dcf = dcf_valuation(target, tax_rate, wacc, terminal_growth=term_growth, projection=projection)
    pf = build_pro_forma(buyer, target, offer, syn, tax_rate, debt_pct, stock_pct, cash_yield,
                         synergy_ramp=synergy_ramp, transaction_fee=transaction_fee,
                         cost_to_achieve=cost_to_achieve, intangible_pct=intangible_pct,
                         intangible_life=intangible_life, financing_fee_pct=financing_fee_pct,
                         financing_life=financing_life)
    verdict = build_verdict(offer, syn, sn, pf, dcf)
    return {"offer": offer, "synergies": syn, "synergy_npv": sn, "dcf": dcf,
            "pf": pf, "verdict": verdict, "wacc": wacc}


def build_scenarios(buyer, target, base_kwargs):
    base = run_case(buyer, target, **base_kwargs)
    bull_kw = dict(base_kwargs)
    bull_kw.update(cost_syn=base_kwargs["cost_syn"] * 1.5, rev_syn=base_kwargs["rev_syn"] * 1.5,
                   syn_margin=min(base_kwargs["syn_margin"] + 0.10, 0.9))
    bear_kw = dict(base_kwargs)
    bear_kw.update(cost_syn=base_kwargs["cost_syn"] * 0.6, rev_syn=base_kwargs["rev_syn"] * 0.6,
                   syn_margin=max(base_kwargs["syn_margin"] - 0.10, 0.0))
    return {"bull": run_case(buyer, target, **bull_kw), "base": base,
            "bear": run_case(buyer, target, **bear_kw)}


# ======================================================================
# Formatting helpers
# ======================================================================

def fmt_pct(x, d=1):
    return "N/A" if x is None else f"{x*100:.{d}f}%"


def fmt_mm(x):
    return "N/A" if x is None else f"${x/1e6:,.0f}mm"


def fmt_x(x, d=2):
    return "N/A" if x is None else f"{x:.{d}f}x"


# ======================================================================
# Excel export
# ======================================================================

_border = Border(*(Side(style="thin"),) * 4)
_hfill = PatternFill("solid", fgColor="1F3864")
_hfont = Font(bold=True, color="FFFFFF")


def _write_table(ws, r0, c0, headers, rows, title=None, num_fmt=None):
    r = r0
    if title:
        ws.cell(r, c0, title).font = Font(bold=True, size=12, color="1F3864")
        r += 1
    for j, h in enumerate(headers):
        cell = ws.cell(r, c0 + j, h)
        cell.font, cell.fill, cell.border = _hfont, _hfill, _border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    r += 1
    for row in rows:
        for j, v in enumerate(row):
            cell = ws.cell(r, c0 + j, v)
            cell.border = _border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if num_fmt and j < len(num_fmt) and num_fmt[j] and isinstance(v, (int, float)):
                cell.number_format = num_fmt[j]
        r += 1
    return r


def _autosize(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 42)


_INPUT_FONT = Font(color="0000CC")            # blue = editable input (banker convention)
_TITLE_FONT = Font(bold=True, size=14, color="1F3864")
_SUBHEAD = Font(bold=True, color="1F3864")


def _sheet_writer(ws, start_row=3):
    """Returns helpers that write label/value rows and remember cell addresses.
    Each write returns (local_ref, global_ref) e.g. ('C5', "'Accretion'!C5")."""
    state = {"r": start_row}

    def row(label, content, fmt=None, blue=False, bold=False):
        r = state["r"]
        lc = ws.cell(r, 2, label)
        if bold:
            lc.font = _SUBHEAD
        c = ws.cell(r, 3, content)
        if fmt:
            c.number_format = fmt
        if blue:
            c.font = _INPUT_FONT
        state["r"] += 1
        return f"C{r}", f"'{ws.title}'!C{r}"

    def head(text):
        ws.cell(state["r"], 2, text).font = _SUBHEAD
        state["r"] += 1

    def blank():
        state["r"] += 1

    return row, head, blank


def build_excel(buyer, target, base, scenarios, comps, meta, inp) -> Workbook:
    """A live, formula-driven, bank-style merger model. Blue cells on 'Assumptions'
    are inputs; every other sheet is built from formulas that reference them, so
    changing an assumption in Excel recalculates the whole model."""
    wb = Workbook()

    # ---------------- Assumptions (the only input sheet) ----------------
    A = {}
    wsa = wb.active
    wsa.title = "Assumptions"
    wsa.cell(1, 2, f"M&A Model — {buyer['name']} acquires {target['name']}").font = _TITLE_FONT
    wsa.cell(2, 2, "Blue cells are inputs. Everything else is formula-linked.").font = Font(italic=True)
    row, head, blank = _sheet_writer(wsa, 4)

    head("Deal terms")
    A["offer_eq"] = row("Offer equity value ($)", inp["offer_equity"], "#,##0", blue=True)[1]
    A["offer_ev"] = row("Offer enterprise value ($)", inp["offer_ev"], "#,##0", blue=True)[1]
    A["debt_pct"] = row("Debt financing (%)", inp["debt_pct"], "0.0%", blue=True)[1]
    A["stock_pct"] = row("Stock financing (%)", inp["stock_pct"], "0.0%", blue=True)[1]
    A["tax"] = row("Tax rate (%)", inp["tax_rate"], "0.0%", blue=True)[1]
    A["cash_yield"] = row("Cash yield / foregone (%)", inp["cash_yield"], "0.0%", blue=True)[1]
    A["cod"] = row("Buyer cost of debt (%)", inp["buyer_cost_of_debt"], "0.0%", blue=True)[1]
    blank()
    head("Fees & purchase accounting")
    A["txn_fee"] = row("Transaction fees ($)", inp["transaction_fee"], "#,##0", blue=True)[1]
    A["cta"] = row("Cost to achieve synergies ($)", inp["cost_to_achieve"], "#,##0", blue=True)[1]
    A["finfee_pct"] = row("Financing fee (% of new debt)", inp["financing_fee_pct"], "0.0%", blue=True)[1]
    A["fin_life"] = row("Financing fee life (yrs)", inp["financing_life"], "0", blue=True)[1]
    A["intang_pct"] = row("Intangibles (% of deal)", inp["intangible_pct"], "0.0%", blue=True)[1]
    A["intang_life"] = row("Intangible amortization life (yrs)", inp["intangible_life"], "0", blue=True)[1]
    blank()
    head("Synergies")
    A["syn_rr"] = row("Annual synergies, after-tax run-rate ($)", inp["synergy_at_runrate"], "#,##0", blue=True)[1]
    A["syn_y1"] = row("Year-1 synergy realization (%)", inp["synergy_y1_factor"], "0%", blue=True)[1]
    blank()
    head("Valuation (DCF)")
    A["wacc"] = row("WACC (%)", inp["wacc"], "0.00%", blue=True)[1]
    A["tg"] = row("Terminal growth (%)", inp["term_growth"], "0.00%", blue=True)[1]
    A["t_rev"] = row("Target revenue ($)", inp["target_revenue"], "#,##0", blue=True)[1]
    A["growth"] = row("Revenue growth (%)", inp["growth"], "0.0%", blue=True)[1]
    A["ebitm"] = row("EBIT margin (%)", inp["ebit_margin"], "0.0%", blue=True)[1]
    A["dapct"] = row("D&A (% of rev)", inp["da_pct"], "0.0%", blue=True)[1]
    A["capexpct"] = row("Capex (% of rev)", inp["capex_pct"], "0.0%", blue=True)[1]
    A["nwcpct"] = row("Δ NWC (% of rev)", inp["nwc_pct"], "0.0%", blue=True)[1]
    blank()
    head("Buyer")
    A["b_ni"] = row("Buyer net income ($)", inp["buyer_ni"], "#,##0", blue=True)[1]
    A["b_sh"] = row("Buyer shares outstanding", inp["buyer_shares"], "#,##0", blue=True)[1]
    A["b_px"] = row("Buyer share price ($)", inp["buyer_price"], "#,##0.00", blue=True)[1]
    A["b_ebd"] = row("Buyer adjusted EBITDA ($)", inp["buyer_adj_ebitda"], "#,##0", blue=True)[1]
    A["b_nd"] = row("Buyer net debt ($)", inp["buyer_net_debt"], "#,##0", blue=True)[1]
    blank()
    head("Target")
    A["t_ni"] = row("Target net income ($)", inp["target_ni"], "#,##0", blue=True)[1]
    A["t_ebd"] = row("Target adjusted EBITDA ($)", inp["target_adj_ebitda"], "#,##0", blue=True)[1]
    A["t_nd"] = row("Target net debt ($)", inp["target_net_debt"], "#,##0", blue=True)[1]
    _autosize(wsa)

    # ---------------- Sources & Uses ----------------
    wsu = wb.create_sheet("Sources_Uses")
    wsu.cell(1, 2, "Sources & Uses").font = _TITLE_FONT
    row, head, blank = _sheet_writer(wsu, 3)
    head("Uses of funds ($)")
    l_pur, _ = row("Purchase of equity", f"={A['offer_eq']}", "#,##0")
    l_txn, _ = row("Transaction fees", f"={A['txn_fee']}", "#,##0")
    l_cta, _ = row("Cost to achieve synergies", f"={A['cta']}", "#,##0")
    l_finf, g_finf = row("Financing fees", f"={A['finfee_pct']}*{A['debt_pct']}*{A['offer_eq']}", "#,##0")
    l_utot, g_utot = row("Total uses", f"={l_pur}+{l_txn}+{l_cta}+{l_finf}", "#,##0", bold=True)
    blank()
    head("Sources of funds ($)")
    l_nd, g_nd = row("New debt", f"={A['debt_pct']}*{A['offer_eq']}", "#,##0")
    l_stk, g_stk = row("Stock issued", f"={A['stock_pct']}*{A['offer_eq']}", "#,##0")
    l_cash, g_cash = row("Cash", f"={l_utot}-{l_nd}-{l_stk}", "#,##0")
    row("Total sources", f"={l_nd}+{l_stk}+{l_cash}", "#,##0", bold=True)
    _autosize(wsu)

    # ---------------- Accretion / Dilution ----------------
    wsc = wb.create_sheet("Accretion")
    wsc.cell(1, 2, "Accretion / Dilution").font = _TITLE_FONT
    row, head, blank = _sheet_writer(wsc, 3)
    head("Pro-forma net income ($)")
    l_bni, _ = row("Buyer net income", f"={A['b_ni']}", "#,##0")
    l_tni, _ = row("Target net income", f"={A['t_ni']}", "#,##0")
    l_syn, _ = row("+ Synergies (Yr-1, after-tax)", f"={A['syn_rr']}*{A['syn_y1']}", "#,##0")
    l_int, _ = row("– New-debt interest (after-tax)", f"=-{g_nd}*{A['cod']}*(1-{A['tax']})", "#,##0")
    l_for, _ = row("– Foregone interest on cash (after-tax)", f"=-{g_cash}*{A['cash_yield']}*(1-{A['tax']})", "#,##0")
    l_ita, _ = row("– Intangible amortization (after-tax)",
                   f"=-IFERROR({A['offer_eq']}*{A['intang_pct']}/{A['intang_life']},0)*(1-{A['tax']})", "#,##0")
    l_ffa, _ = row("– Financing-fee amortization (after-tax)",
                   f"=-IFERROR({g_finf}/{A['fin_life']},0)*(1-{A['tax']})", "#,##0")
    l_pfni, g_pfni = row("Pro-forma net income", f"={l_bni}+{l_tni}+{l_syn}+{l_int}+{l_for}+{l_ita}+{l_ffa}",
                         "#,##0", bold=True)
    blank()
    head("Per-share")
    l_nsh, _ = row("New shares issued", f"=IFERROR({g_stk}/{A['b_px']},0)", "#,##0")
    l_pfsh, _ = row("Pro-forma shares", f"={A['b_sh']}+{l_nsh}", "#,##0")
    l_beps, _ = row("Buyer standalone EPS", f"=IFERROR({A['b_ni']}/{A['b_sh']},0)", "#,##0.00")
    l_peps, g_peps = row("Pro-forma EPS", f"=IFERROR({l_pfni}/{l_pfsh},0)", "#,##0.00")
    l_acc, g_acc = row("Accretion / (dilution)", f"=IFERROR({l_peps}/{l_beps}-1,0)", "0.0%", bold=True)
    blank()
    head("Leverage")
    l_pnd, _ = row("Pro-forma net debt", f"={A['b_nd']}+{A['t_nd']}+{g_nd}", "#,##0")
    l_peb, _ = row("Pro-forma EBITDA (incl. synergies)",
                   f"={A['b_ebd']}+{A['t_ebd']}+IFERROR({A['syn_rr']}/(1-{A['tax']}),0)", "#,##0")
    l_lev, g_lev = row("Pro-forma leverage (x)", f"=IFERROR({l_pnd}/{l_peb},0)", "0.00x", bold=True)
    _autosize(wsc)

    # ---------------- DCF ----------------
    wsd = wb.create_sheet("DCF")
    wsd.cell(1, 2, "DCF — Target (unlevered FCF)").font = _TITLE_FONT
    row, head, blank = _sheet_writer(wsd, 3)
    head("Forecast")
    pv_locals, prev_rev, last_fcf = [], None, None
    for i in range(1, 6):
        if i == 1:
            rev_f = f"={A['t_rev']}*(1+{A['growth']})"
        else:
            rev_f = f"={prev_rev}*(1+{A['growth']})"
        l_rev, _ = row(f"Y{i} revenue", rev_f, "#,##0")
        l_fcf, _ = row(f"Y{i} unlevered FCF",
                       f"={l_rev}*({A['ebitm']}*(1-{A['tax']})+{A['dapct']}-{A['capexpct']}-{A['nwcpct']})", "#,##0")
        l_pv, _ = row(f"Y{i} PV of FCF", f"={l_fcf}/(1+{A['wacc']})^{i}", "#,##0")
        pv_locals.append(l_pv)
        prev_rev, last_fcf = l_rev, l_fcf
    blank()
    head("Valuation")
    l_tv, _ = row("Terminal value", f"=IFERROR({last_fcf}*(1+{A['tg']})/({A['wacc']}-{A['tg']}),0)", "#,##0")
    l_tvpv, _ = row("PV of terminal value", f"={l_tv}/(1+{A['wacc']})^5", "#,##0")
    l_ev, g_ev = row("Enterprise value", f"={'+'.join(pv_locals)}+{l_tvpv}", "#,##0", bold=True)
    l_deq, g_deq = row("Equity value (EV − net debt)", f"={l_ev}-{A['t_nd']}", "#,##0", bold=True)
    _autosize(wsd)

    # ---------------- Summary (links everything) ----------------
    wss = wb.create_sheet("Summary")
    wb.move_sheet("Summary", -(len(wb.sheetnames) - 1))   # put Summary first
    wss.cell(1, 2, "Deal Summary").font = _TITLE_FONT
    wss.cell(2, 2, f"Generated {meta['generated']} · buyer data: {buyer['source']} · target data: {target['source']}")
    row, head, blank = _sheet_writer(wss, 4)
    row("Offer equity value", f"={A['offer_eq']}", "#,##0")
    row("Offer enterprise value", f"={A['offer_ev']}", "#,##0")
    row("EPS accretion / (dilution)", f"={g_acc}", "0.0%", bold=True)
    row("Pro-forma EPS", f"={g_peps}", "#,##0.00")
    row("Pro-forma leverage", f"={g_lev}", "0.00x", bold=True)
    row("DCF enterprise value", f"={g_ev}", "#,##0")
    row("DCF equity value", f"={g_deq}", "#,##0")
    row("DCF vs offer equity", f"={g_deq}-{A['offer_eq']}", "#,##0")
    row("Synergy NPV (after-tax)", base["synergy_npv"]["synergy_npv"], "#,##0")
    row("Screening verdict", base["verdict"]["verdict"], bold=True)
    _autosize(wss)

    # ---------------- Scenarios (static reference) ----------------
    ws2 = wb.create_sheet("Scenarios")
    rows = [[lbl.capitalize(), c["pf"]["accretion_pct"], c["pf"]["pf_leverage"],
             c["dcf"]["equity_value"], c["verdict"]["verdict"]]
            for lbl, c in [("bull", scenarios["bull"]), ("base", scenarios["base"]),
                           ("bear", scenarios["bear"])]]
    _write_table(ws2, 2, 2, ["Case", "Accretion", "PF Leverage", "DCF Equity", "Verdict"],
                 rows, "Scenario Analysis (synergy sensitivity)", num_fmt=[None, "0.0%", "0.00x", "#,##0", None])
    _autosize(ws2)

    if comps and comps.get("n_peers"):
        ws3 = wb.create_sheet("Comps")
        rows = [["Peers used", comps["n_peers"]],
                ["Median EV/EBITDA", comps["median_ev_ebitda"]],
                ["Median EV/Revenue", comps["median_ev_rev"]],
                ["Implied EV (EBITDA)", comps["implied_ev_from_ebitda"]],
                ["Implied equity (EBITDA)", comps["implied_equity_from_ebitda"]],
                ["Implied EV (Revenue)", comps["implied_ev_from_rev"]]]
        _write_table(ws3, 2, 2, ["Comps metric", "Value"], rows, "Trading Comps",
                     num_fmt=[None, "#,##0.00"])
        _autosize(ws3)

    return wb


# ======================================================================
# Streamlit UI
# ======================================================================

# widget-key defaults (also the schema for save/load)
DEFAULTS = {
    # deal assumptions
    "offer_basis": "Premium to share price", "premium_pct": 30.0,
    "ev_multiple": 10.0, "ev_abs_mm": 0.0,
    "debt_pct": 60.0, "stock_pct": 0.0, "tax_rate": 25.0, "cash_yield": 4.0,
    "rf": 4.0, "erp": 5.0, "term_growth": 2.5, "wacc_override": 0.0,
    "cost_syn": 5.0, "rev_syn": 3.0, "syn_margin": 30.0,
    "peers": "", "run_comps": False,
    # merger inputs (real-world)
    "transaction_fee_mm": 0.0, "cost_to_achieve_mm": 0.0,
    "intangible_pct": 0.0, "intangible_life": 15.0, "financing_life": 7.0,
    "financing_fee_pct": 0.0,
    "syn_ramp_y1": 25.0, "syn_ramp_y2": 60.0, "syn_ramp_y3": 100.0,
    # target multi-year forecast (optional)
    "use_projection": False,
}
for i in range(1, 6):
    DEFAULTS[f"fc_rev_{i}"] = 0.0
    DEFAULTS[f"fc_ebitda_{i}"] = 0.0
for role in ("buyer", "target"):
    DEFAULTS.update({
        f"{role}_source": "Manual entry", f"{role}_ticker": "",
        f"{role}_name": role.capitalize(), f"{role}_sector": "", f"{role}_industry": "",
        f"{role}_price": 0.0, f"{role}_shares_mm": 0.0, f"{role}_rev_mm": 0.0,
        f"{role}_ebitda_mm": 0.0, f"{role}_da_mm": 0.0,
        f"{role}_capex_mm": 0.0, f"{role}_nwc_mm": 0.0, f"{role}_ni_mm": 0.0,
        f"{role}_int_mm": 0.0, f"{role}_debt_mm": 0.0, f"{role}_cash_mm": 0.0,
        f"{role}_beta": 0.0, f"{role}_growth_pct": 0.0, f"{role}_edgar_price": 0.0,
        # itemized add-backs ($mm)
        f"{role}_ab_owner": 0.0, f"{role}_ab_onetime": 0.0,
        f"{role}_ab_personal": 0.0, f"{role}_ab_other": 0.0,
    })


def _mm(v):
    return v * 1e6 if v else None


def resolve_snapshot(role: str) -> dict:
    ss = st.session_state
    src = ss[f"{role}_source"]
    if src == "SEC EDGAR":
        return fetch_from_edgar(ss[f"{role}_ticker"], price=ss[f"{role}_edgar_price"] or None)
    if src == "Yahoo Finance":
        return fetch_from_yahoo(ss[f"{role}_ticker"])
    # Manual entry (values entered in $mm / millions of shares)
    add_backs_mm = (ss[f"{role}_ab_owner"] + ss[f"{role}_ab_onetime"]
                    + ss[f"{role}_ab_personal"] + ss[f"{role}_ab_other"])
    return finalize_snapshot({
        "ticker": ss[f"{role}_ticker"] or role.upper(),
        "name": ss[f"{role}_name"], "sector": ss[f"{role}_sector"],
        "industry": ss[f"{role}_industry"], "price": ss[f"{role}_price"] or None,
        "shares_out": _mm(ss[f"{role}_shares_mm"]), "revenue": _mm(ss[f"{role}_rev_mm"]),
        "ebitda": _mm(ss[f"{role}_ebitda_mm"]), "add_backs": (add_backs_mm * 1e6) if add_backs_mm else 0.0,
        "da": _mm(ss[f"{role}_da_mm"]), "capex": _mm(ss[f"{role}_capex_mm"]),
        "nwc_change": _mm(ss[f"{role}_nwc_mm"]), "net_income": _mm(ss[f"{role}_ni_mm"]),
        "interest_expense": _mm(ss[f"{role}_int_mm"]), "total_debt": _mm(ss[f"{role}_debt_mm"]) or 0.0,
        "cash": _mm(ss[f"{role}_cash_mm"]) or 0.0, "beta": ss[f"{role}_beta"] or None,
        "revenue_growth": (ss[f"{role}_growth_pct"] / 100) if ss[f"{role}_growth_pct"] else None,
        "source": "Manual entry",
    })


def company_input_ui(role: str):
    ss = st.session_state
    st.markdown(f"#### {role.capitalize()}")
    st.selectbox("Data source", ["Manual entry", "SEC EDGAR", "Yahoo Finance"],
                 key=f"{role}_source")
    src = ss[f"{role}_source"]

    if src in ("SEC EDGAR", "Yahoo Finance"):
        st.text_input("Ticker", key=f"{role}_ticker",
                      placeholder="e.g. XOM").upper() if False else st.text_input("Ticker", key=f"{role}_ticker")
        if src == "SEC EDGAR":
            st.number_input("Current share price ($) — not in filings, enter for market cap",
                            min_value=0.0, step=1.0, key=f"{role}_edgar_price")
            st.caption("Pulls audited revenue, EBITDA (built from operating income + D&A), "
                       "D&A, capex, debt, cash and shares from the latest 10-K.")
        return

    # Manual entry
    st.text_input("Company name", key=f"{role}_name")
    c1, c2 = st.columns(2)
    c1.text_input("Ticker (optional)", key=f"{role}_ticker")
    c2.text_input("Sector", key=f"{role}_sector")
    st.caption("Enter figures in **$ millions** (shares in millions). Price is per-share. "
               "You can type any number directly — the +/- buttons are just for small nudges.")
    c1, c2 = st.columns(2)
    c1.number_input("Revenue", min_value=0.0, step=1.0, key=f"{role}_rev_mm")
    c2.number_input("EBITDA (reported)", step=1.0, key=f"{role}_ebitda_mm")

    with st.expander("Add-backs → Adjusted EBITDA"):
        st.caption("Itemize normalizations so Adjusted EBITDA is auditable ($mm).")
        a1, a2 = st.columns(2)
        a1.number_input("Owner's excess comp", step=1.0, key=f"{role}_ab_owner")
        a2.number_input("One-time / non-recurring", step=1.0, key=f"{role}_ab_onetime")
        a1.number_input("Personal expenses", step=1.0, key=f"{role}_ab_personal")
        a2.number_input("Other", step=1.0, key=f"{role}_ab_other")
        total_ab = (ss[f"{role}_ab_owner"] + ss[f"{role}_ab_onetime"]
                    + ss[f"{role}_ab_personal"] + ss[f"{role}_ab_other"])
        adj = (ss[f"{role}_ebitda_mm"] or 0) + total_ab
        st.markdown(f"**Total add-backs: ${total_ab:,.0f}mm → Adjusted EBITDA: ${adj:,.0f}mm**")

    c1, c2, c3 = st.columns(3)
    c1.number_input("D&A", min_value=0.0, step=1.0, key=f"{role}_da_mm")
    c2.number_input("Capex", min_value=0.0, step=1.0, key=f"{role}_capex_mm")
    c3.number_input("Δ Net working capital", step=1.0, key=f"{role}_nwc_mm")
    c1, c2, c3 = st.columns(3)
    c1.number_input("Net income", step=1.0, key=f"{role}_ni_mm")
    c2.number_input("Interest expense", min_value=0.0, step=1.0, key=f"{role}_int_mm")
    c3.number_input("Rev. growth (%)", step=1.0, key=f"{role}_growth_pct")
    c1, c2, c3 = st.columns(3)
    c1.number_input("Total debt", min_value=0.0, step=1.0, key=f"{role}_debt_mm")
    c2.number_input("Cash", min_value=0.0, step=1.0, key=f"{role}_cash_mm")
    c3.number_input("Beta (optional)", min_value=0.0, step=0.1, key=f"{role}_beta")
    c1, c2 = st.columns(2)
    c1.number_input("Share price ($)", min_value=0.0, step=1.0, key=f"{role}_price")
    c2.number_input("Shares out (millions)", min_value=0.0, step=1.0, key=f"{role}_shares_mm")


def fetch_peers(peer_str: str) -> list:
    out = []
    for tk in [p.strip().upper() for p in peer_str.split(",") if p.strip()]:
        try:
            out.append(fetch_from_yahoo(tk))
        except Exception:
            continue
    return out


def target_projection_ui():
    """Optional multi-year forecast for the TARGET (drives the DCF when filled in)."""
    ss = st.session_state
    with st.expander("Advanced (optional): 5-year forecast for the target"):
        st.caption("Fill in a 5-year revenue & EBITDA forecast to drive the DCF directly "
                   "instead of a single flat growth rate. Leave off for the quick single-year DCF. ($mm)")
        st.checkbox("Use my 5-year forecast in the DCF", key="use_projection")
        h = st.columns(6)
        h[0].markdown("**Year**")
        for i in range(1, 6):
            h[i].markdown(f"**Y{i}**")
        r = st.columns(6)
        r[0].markdown("Revenue")
        for i in range(1, 6):
            r[i].number_input(f"rev{i}", min_value=0.0, step=1.0, key=f"fc_rev_{i}",
                              label_visibility="collapsed")
        e = st.columns(6)
        e[0].markdown("EBITDA")
        for i in range(1, 6):
            e[i].number_input(f"ebitda{i}", step=1.0, key=f"fc_ebitda_{i}",
                              label_visibility="collapsed")


def build_projection():
    """Return a 5-year projection list from session state, or None if not used/incomplete."""
    ss = st.session_state
    if not ss.get("use_projection"):
        return None
    proj = []
    for i in range(1, 6):
        rev = ss.get(f"fc_rev_{i}", 0.0)
        if not rev:
            return None                       # incomplete -> fall back to flat-growth DCF
        eb = ss.get(f"fc_ebitda_{i}", 0.0)
        proj.append({"revenue": rev * 1e6, "ebitda": (eb * 1e6) if eb else None})
    return proj


def main():
    st.set_page_config(page_title="M&A Deal Screener", layout="wide")
    for k, val in DEFAULTS.items():
        st.session_state.setdefault(k, val)

    st.title("M&A Deal Screener")
    st.caption("Public **or** private targets · manual entry, SEC EDGAR (free filings) or Yahoo · "
               "unlevered-FCF DCF · accretion/dilution · trading comps · Excel export")

    # ---------- Sidebar: load, assumptions, save ----------
    with st.sidebar:
        st.header("Load a saved deal")
        up = st.file_uploader("Load deal (.json)", type="json", label_visibility="collapsed")
        if up is not None:
            sig = up.name + str(up.size)
            if st.session_state.get("_loaded_sig") != sig:
                try:
                    data = json.load(up)
                    for k, val in data.items():
                        if k in DEFAULTS:
                            st.session_state[k] = val
                    st.session_state["_loaded_sig"] = sig
                    st.success("Deal loaded.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Could not load: {e}")

        st.header("Deal assumptions")
        st.selectbox("Offer basis",
                     ["Premium to share price", "EV / EBITDA multiple", "Enterprise value ($mm)"],
                     key="offer_basis",
                     help="Use a premium for public targets; a multiple or absolute EV for private targets.")
        basis = st.session_state["offer_basis"]
        if basis == "Premium to share price":
            st.number_input("Offer premium (%)", step=5.0, key="premium_pct")
        elif basis == "EV / EBITDA multiple":
            st.number_input("Purchase EV / EBITDA (x)", min_value=0.0, step=0.5, key="ev_multiple")
        else:
            st.number_input("Purchase enterprise value ($mm)", min_value=0.0, step=50.0, key="ev_abs_mm")

        st.number_input("Debt financing (%)", 0.0, 100.0, step=5.0, key="debt_pct")
        st.number_input("Stock financing (%)", 0.0, 100.0, step=5.0, key="stock_pct")
        st.caption("Remainder is funded with cash.")
        st.number_input("Tax rate (%)", 0.0, 60.0, step=1.0, key="tax_rate")
        st.number_input("Cash yield — foregone (%)", 0.0, 15.0, step=0.5, key="cash_yield")

        with st.expander("Valuation / WACC assumptions"):
            st.number_input("Risk-free rate (%)", 0.0, 10.0, step=0.25, key="rf")
            st.number_input("Equity risk premium (%)", 0.0, 12.0, step=0.25, key="erp")
            st.number_input("Terminal growth (%)", 0.0, 5.0, step=0.25, key="term_growth")
            st.number_input("WACC override (%) — 0 = auto", 0.0, 30.0, step=0.5, key="wacc_override")

        with st.expander("Synergy assumptions"):
            st.number_input("Cost synergies (% of target rev)", 0.0, 50.0, step=0.5, key="cost_syn")
            st.number_input("Revenue synergies (% of target rev)", 0.0, 50.0, step=0.5, key="rev_syn")
            st.number_input("Revenue-synergy margin (%)", 0.0, 100.0, step=5.0, key="syn_margin")
            st.caption("Synergy ramp — % of run-rate realized each year:")
            s1, s2, s3 = st.columns(3)
            s1.number_input("Y1 (%)", 0.0, 100.0, step=5.0, key="syn_ramp_y1")
            s2.number_input("Y2 (%)", 0.0, 100.0, step=5.0, key="syn_ramp_y2")
            s3.number_input("Y3+ (%)", 0.0, 100.0, step=5.0, key="syn_ramp_y3")

        with st.expander("Merger inputs (fees & purchase accounting)"):
            st.number_input("Transaction fees ($mm)", 0.0, step=1.0, key="transaction_fee_mm",
                            help="Advisory & legal fees — one-time, funded with cash.")
            st.number_input("Cost to achieve synergies ($mm)", 0.0, step=1.0, key="cost_to_achieve_mm",
                            help="One-time integration/restructuring cost.")
            st.number_input("Financing fee (% of new debt)", 0.0, 5.0, step=0.25, key="financing_fee_pct")
            st.number_input("Financing fee life (yrs)", 1.0, 30.0, step=1.0, key="financing_life")
            st.number_input("Intangibles created (% of deal value)", 0.0, 100.0, step=5.0,
                            key="intangible_pct",
                            help="Portion of the price allocated to amortizable intangibles (step-up).")
            st.number_input("Intangible amortization life (yrs)", 1.0, 40.0, step=1.0, key="intangible_life")

        with st.expander("Trading comps (optional)"):
            st.checkbox("Run comps", key="run_comps")
            st.text_input("Peer tickers (comma-separated)", key="peers",
                          placeholder="CVX, SHEL, BP, COP")

        run = st.button("Run deal analysis", type="primary", use_container_width=True)

        # Save current inputs
        save_blob = json.dumps({k: st.session_state[k] for k in DEFAULTS}, indent=2)
        st.download_button("💾 Save deal (.json)", save_blob,
                           file_name="ma_deal.json", mime="application/json",
                           use_container_width=True)

    # ---------- Company inputs ----------
    cbuy, ctar = st.columns(2)
    with cbuy:
        company_input_ui("buyer")
    with ctar:
        company_input_ui("target")
        target_projection_ui()

    with st.expander("⚠️ Assumptions & limitations (read before relying on output)"):
        st.markdown(
            "- **Screening tool, not a fairness opinion.** Simplifies a full three-statement merger model.\n"
            "- **Add-backs / normalization are the user's judgment** — the tool does not verify them.\n"
            "- **DCF** uses single-stage growth (or your 5-year forecast) and a maintenance-capex≈D&A proxy when capex/D&A are blank.\n"
            "- **EDGAR** covers US public filers only and does not contain a share price (enter it manually).\n"
            "- **Purchase accounting** is simplified (a single intangible pool amortized straight-line; no goodwill/deferred-tax detail).\n"
            "- Data is from free sources (EDGAR/Yahoo) or manual entry — not Bloomberg/Capital IQ grade.\n"
            "- The verdict is an **illustrative flag**, never an investment recommendation."
        )

    if not run:
        st.info("Set up the buyer and target on the left/right, choose your assumptions, then **Run deal analysis**.")
        return

    # ---------- Resolve data ----------
    try:
        with st.spinner("Fetching / building company data..."):
            buyer = resolve_snapshot("buyer")
            target = resolve_snapshot("target")
    except Exception as e:
        st.error(f"Could not build company data: {e}")
        return

    for label, snap in [("Buyer", buyer), ("Target", target)]:
        for w in snapshot_completeness(snap):
            st.warning(f"{label}: {w}")

    # ---------- Assemble offer kwargs ----------
    ss = st.session_state
    if basis == "Premium to share price":
        offer_kwargs = {"mode": "premium", "premium": ss["premium_pct"] / 100}
    elif basis == "EV / EBITDA multiple":
        offer_kwargs = {"mode": "multiple", "ev_multiple": ss["ev_multiple"]}
    else:
        offer_kwargs = {"mode": "ev_abs", "ev_abs": ss["ev_abs_mm"] * 1e6}

    debt_pct, stock_pct = ss["debt_pct"] / 100, ss["stock_pct"] / 100
    if debt_pct + stock_pct > 1.0:
        stock_pct = max(0.0, 1.0 - debt_pct)
        st.warning("Debt + stock financing exceeded 100%; stock reduced so the mix sums to 100%.")

    projection = build_projection()
    case_kwargs = dict(
        offer_kwargs=offer_kwargs, debt_pct=debt_pct, stock_pct=stock_pct,
        tax_rate=ss["tax_rate"] / 100, cost_syn=ss["cost_syn"] / 100, rev_syn=ss["rev_syn"] / 100,
        syn_margin=ss["syn_margin"] / 100, cash_yield=ss["cash_yield"] / 100,
        rf=ss["rf"] / 100, erp=ss["erp"] / 100, term_growth=ss["term_growth"] / 100,
        wacc_override=(ss["wacc_override"] / 100) if ss["wacc_override"] else None,
        synergy_ramp=(ss["syn_ramp_y1"] / 100, ss["syn_ramp_y2"] / 100, ss["syn_ramp_y3"] / 100),
        transaction_fee=ss["transaction_fee_mm"] * 1e6, cost_to_achieve=ss["cost_to_achieve_mm"] * 1e6,
        intangible_pct=ss["intangible_pct"] / 100, intangible_life=ss["intangible_life"] or 15,
        financing_fee_pct=ss["financing_fee_pct"] / 100, financing_life=ss["financing_life"] or 7,
        projection=projection,
    )

    try:
        base = run_case(buyer, target, **case_kwargs)
        scenarios = build_scenarios(buyer, target, case_kwargs)
    except Exception as e:
        st.error(f"Model error: {e}")
        return

    comps = None
    if ss["run_comps"] and ss["peers"].strip():
        peers = fetch_peers(ss["peers"])
        if peers:
            comps = comps_valuation(target, peers)

    v, offer, pf, dcf = base["verdict"], base["offer"], base["pf"], base["dcf"]

    # ---------- Results ----------
    tab_ov, tab_dcf, tab_comps, tab_scn, tab_exp = st.tabs(
        ["📊 Overview", "💵 DCF", "📈 Comps", "🔀 Scenarios", "⬇️ Export"])

    with tab_ov:
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("EPS accretion (Yr-1)", fmt_pct(pf["accretion_pct"]),
                  help="Year-1, with synergies phased in.")
        k2.metric("Accretion (run-rate)", fmt_pct(pf["accretion_pct_runrate"]),
                  help="With synergies fully realized.")
        k3.metric("PF leverage", fmt_x(pf["pf_leverage"]))
        k4.metric("Implied EV/EBITDA", fmt_x(offer["implied_ev_ebitda"]))
        st.subheader(v["verdict"])
        c1, c2, c3 = st.columns(3)
        c1.write(f"**Offer EV:** {fmt_mm(offer['ev_offer'])}")
        c1.write(f"**Offer equity:** {fmt_mm(offer['equity_value_offer'])}")
        c1.write(f"**Implied premium:** {fmt_pct(offer['premium_pct'], 0)}")
        c2.write(f"**Synergies (a/t run-rate):** {fmt_mm(pf['synergy_after_tax'])}")
        c2.write(f"**Synergy NPV (net of cost-to-achieve):** {fmt_mm(v['synergy_npv'])}")
        c2.write(f"**Synergy realized Yr-1:** {fmt_mm(pf['synergy_after_tax_y1'])}")
        c3.write(f"**DCF equity value:** {fmt_mm(dcf['equity_value'])}")
        c3.write(f"**DCF − offer equity:** {fmt_mm(v['dcf_vs_offer'])}")
        c3.write(f"**WACC:** {fmt_pct(base['wacc'])}")

        with st.expander("Deal build — fees, financing & purchase accounting"):
            b1, b2, b3 = st.columns(3)
            b1.write(f"**New debt:** {fmt_mm(pf['new_debt'])}")
            b1.write(f"**Equity issued:** {fmt_mm(pf['equity_issued_value'])}")
            b1.write(f"**Cash used:** {fmt_mm(pf['cash_used'])}")
            b2.write(f"**Transaction fees:** {fmt_mm(pf['transaction_fee'])}")
            b2.write(f"**Cost to achieve:** {fmt_mm(pf['cost_to_achieve'])}")
            b2.write(f"**Financing fees:** {fmt_mm(pf['financing_fee'])}")
            b3.write(f"**Intangibles created:** {fmt_mm(pf['intangibles'])}")
            b3.write(f"**Intangible amort (a/t):** {fmt_mm(pf['intangible_amort_at'])}")
            b3.write(f"**Fin-fee amort (a/t):** {fmt_mm(pf['fin_fee_amort_at'])}")
        st.caption(f"Buyer source: {buyer['source']} · Target source: {target['source']}")

    with tab_dcf:
        st.subheader("Discounted cash flow — target (unlevered FCF)")
        if dcf["fcf_list"]:
            if projection:
                st.success("Using your 5-year forecast for the DCF.")
            elif dcf["assumed_da"]:
                st.info("D&A wasn't provided, so capex≈D&A and a D&A proxy (~4% of revenue) were assumed.")
            df = pd.DataFrame({
                "Year": [f"Y{i}" for i in range(1, len(dcf["fcf_list"]) + 1)],
                "Revenue ($mm)": [round(r / 1e6, 1) for r in (dcf.get("revenue_proj") or [])] or [None] * len(dcf["fcf_list"]),
                "Unlevered FCF ($mm)": [round(f / 1e6, 1) for f in dcf["fcf_list"]],
            })
            st.dataframe(df, use_container_width=True, hide_index=True)
            c1, c2, c3 = st.columns(3)
            c1.metric("Enterprise value", fmt_mm(dcf["enterprise_value"]))
            c2.metric("Equity value", fmt_mm(dcf["equity_value"]))
            c3.metric("Terminal value", fmt_mm(dcf["terminal_value"]))
        else:
            st.warning("DCF needs revenue, EBITDA and a valid WACC above the terminal growth rate.")

    with tab_comps:
        st.subheader("Trading comparables")
        if comps:
            c1, c2 = st.columns(2)
            c1.metric("Median EV/EBITDA", fmt_x(comps["median_ev_ebitda"]))
            c2.metric("Median EV/Revenue", fmt_x(comps["median_ev_rev"]))
            st.write(f"Implied EV (EBITDA basis): **{fmt_mm(comps['implied_ev_from_ebitda'])}** · "
                     f"implied equity: **{fmt_mm(comps['implied_equity_from_ebitda'])}**")
            st.write(f"Implied EV (Revenue basis): **{fmt_mm(comps['implied_ev_from_rev'])}**")
            st.caption(f"Based on {comps['n_peers']} peer(s) pulled from Yahoo Finance.")
        else:
            st.info("Enable **Run comps** in the sidebar and enter peer tickers to value the target off "
                    "public trading multiples.")

    with tab_scn:
        st.subheader("Scenario analysis (Bull / Base / Bear synergies)")
        rows = []
        for lbl in ["bull", "base", "bear"]:
            c = scenarios[lbl]
            rows.append({"Case": lbl.capitalize(),
                         "Accretion": fmt_pct(c["verdict"]["accretion_pct"]),
                         "PF Leverage": fmt_x(c["pf"]["pf_leverage"]),
                         "DCF Equity": fmt_mm(c["dcf"]["equity_value"]),
                         "Verdict": c["verdict"]["verdict"]})
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    with tab_exp:
        st.subheader("Export — live, formula-driven Excel model")
        st.caption("Bank-style workbook: blue cells on the **Assumptions** tab are inputs; "
                   "Sources & Uses, Accretion, DCF and Summary are built from formulas that "
                   "recalculate when you change an assumption in Excel.")
        meta = {"generated": datetime.now().strftime("%Y-%m-%d %H:%M")}

        tgt_rev = target.get("revenue") or 0
        tgt_adj = target.get("adjusted_ebitda") or 0
        _da = target.get("da") if target.get("da") is not None else (tgt_rev * 0.04)
        _capex = target.get("capex") if target.get("capex") is not None else _da
        _nwc = target.get("nwc_change") or 0
        inp = {
            "offer_equity": offer["equity_value_offer"] or 0,
            "offer_ev": offer["ev_offer"] or 0,
            "debt_pct": debt_pct, "stock_pct": stock_pct,
            "transaction_fee": ss["transaction_fee_mm"] * 1e6,
            "cost_to_achieve": ss["cost_to_achieve_mm"] * 1e6,
            "financing_fee_pct": ss["financing_fee_pct"] / 100,
            "financing_life": ss["financing_life"] or 7,
            "tax_rate": ss["tax_rate"] / 100, "cash_yield": ss["cash_yield"] / 100,
            "buyer_cost_of_debt": estimate_cost_of_debt(buyer),
            "intangible_pct": ss["intangible_pct"] / 100, "intangible_life": ss["intangible_life"] or 15,
            "wacc": base["wacc"], "term_growth": ss["term_growth"] / 100,
            "synergy_at_runrate": base["synergies"]["annual_synergy_after_tax"],
            "synergy_y1_factor": ss["syn_ramp_y1"] / 100,
            "target_revenue": tgt_rev,
            "growth": target.get("revenue_growth") if target.get("revenue_growth") is not None else 0.05,
            "ebit_margin": ((tgt_adj - _da) / tgt_rev) if tgt_rev else 0.0,
            "da_pct": (_da / tgt_rev) if tgt_rev else 0.0,
            "capex_pct": (_capex / tgt_rev) if tgt_rev else 0.0,
            "nwc_pct": (_nwc / tgt_rev) if tgt_rev else 0.0,
            "buyer_ni": buyer.get("net_income") or 0, "buyer_shares": buyer.get("shares_out") or 0,
            "buyer_price": buyer.get("price") or 0, "buyer_adj_ebitda": buyer.get("adjusted_ebitda") or 0,
            "buyer_net_debt": buyer.get("net_debt") or 0,
            "target_ni": target.get("net_income") or 0, "target_adj_ebitda": tgt_adj,
            "target_net_debt": target.get("net_debt") or 0,
        }
        wb = build_excel(buyer, target, base, scenarios, comps, meta, inp)
        buf = io.BytesIO()
        wb.save(buf)
        st.download_button("⬇️ Download Excel model", buf.getvalue(),
                           file_name=f"deal_{buyer.get('ticker') or 'buyer'}_{target.get('ticker') or 'target'}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)


if __name__ == "__main__":
    main()

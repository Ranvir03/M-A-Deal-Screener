import sys
import os
import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton,
    QGridLayout, QMessageBox, QFileDialog
)
from PyQt5.QtCore import Qt

# =========================
# Data Layer
# =========================

class DataProvider:
    def get_company_snapshot(self, identifier: str) -> dict:
        raise NotImplementedError


class YahooDataProvider(DataProvider):
    def get_company_snapshot(self, ticker: str) -> dict:
        t = yf.Ticker(ticker)
        info = t.info or {}

        def g(key, default=None):
            v = info.get(key, default)
            return v if v is not None else default

        name = g("longName", ticker)
        sector = g("sector", "N/A")
        industry = g("industry", "N/A")

        market_cap = g("marketCap")
        enterprise_value = g("enterpriseValue")
        revenue = g("totalRevenue")
        ebitda = g("ebitda")
        net_income = g("netIncomeToCommon") or g("netIncome")
        shares_out = g("sharesOutstanding")
        price = g("currentPrice") or g("regularMarketPrice")

        revenue_growth = g("revenueGrowth")
        profit_margin = g("profitMargins")
        op_margin = g("operatingMargins")
        gross_margin = g("grossMargins")

        total_debt = g("totalDebt", 0) or 0
        cash = g("totalCash", 0) or 0
        interest_expense = g("interestExpense")
        beta = g("beta")

        ev_rev = enterprise_value / revenue if enterprise_value and revenue else None
        ev_ebitda = enterprise_value / ebitda if enterprise_value and ebitda else None
        net_debt = total_debt - cash if total_debt is not None and cash is not None else None
        debt_ev = net_debt / enterprise_value if enterprise_value and net_debt is not None else None

        snapshot = {
            "ticker": ticker,
            "name": name,
            "sector": sector,
            "industry": industry,
            "market_cap": market_cap,
            "enterprise_value": enterprise_value,
            "revenue": revenue,
            "ebitda": ebitda,
            "net_income": net_income,
            "shares_out": shares_out,
            "price": price,
            "ev_rev": ev_rev,
            "ev_ebitda": ev_ebitda,
            "revenue_growth": revenue_growth,
            "profit_margin": profit_margin,
            "op_margin": op_margin,
            "gross_margin": gross_margin,
            "total_debt": total_debt,
            "cash": cash,
            "net_debt": net_debt,
            "debt_ev": debt_ev,
            "interest_expense": interest_expense,
            "beta": beta,
        }

        if market_cap is None or enterprise_value is None or shares_out is None or price is None:
            raise ValueError(f"Insufficient data for {ticker}")

        return snapshot

# =========================
# Helpers
# =========================

def pct_to_float(s, default):
    if s is None or s == "":
        return default / 100.0
    s = str(s).replace("%", "").strip()
    try:
        return float(s) / 100.0
    except:
        return default / 100.0


def safe_float(s, default):
    if s is None or s == "":
        return default
    s = str(s).replace(",", "").strip()
    try:
        return float(s)
    except:
        return default


def fmt_pct(x):
    if x is None:
        return "N/A"
    return f"{x*100:.1f}%"

# =========================
# WACC & DCF
# =========================

def estimate_cost_of_debt(company: dict, fallback_rate: float = 0.05) -> float:
    interest = company["interest_expense"]
    debt = company["total_debt"]
    if interest and debt and debt > 0:
        rate = abs(interest) / debt
        if 0.0 < rate < 0.2:
            return rate
    return fallback_rate


def estimate_cost_of_equity(company: dict,
                            risk_free: float = 0.04,
                            market_premium: float = 0.05) -> float:
    beta = company.get("beta")
    if beta is None:
        beta = 1.0
    return risk_free + beta * market_premium


def estimate_wacc(company: dict,
                  tax_rate: float,
                  risk_free: float = 0.04,
                  market_premium: float = 0.05) -> float:
    ke = estimate_cost_of_equity(company, risk_free, market_premium)
    kd = estimate_cost_of_debt(company)
    debt = company["total_debt"] or 0
    equity = company["market_cap"] or 0
    total = debt + equity
    if total <= 0:
        return ke
    wd = debt / total
    we = equity / total
    return we * ke + wd * kd * (1 - tax_rate)


def basic_dcf_target(target: dict,
                     tax_rate: float,
                     wacc: float,
                     years: int = 5,
                     terminal_growth: float = 0.02,
                     fcf_conversion: float = 0.6) -> dict:
    revenue = target["revenue"]
    ebitda = target["ebitda"]
    if revenue is None or ebitda is None or wacc <= 0:
        return {"dcf_value": None, "fcf_list": None}

    growth = target["revenue_growth"] or 0.05
    ebitda_margin = ebitda / revenue if revenue else 0.2

    fcf_list = []
    rev = revenue
    for _ in range(years):
        rev = rev * (1 + growth)
        ebitda_year = rev * ebitda_margin
        ebit = ebitda_year
        fcf = ebit * (1 - tax_rate) * fcf_conversion
        fcf_list.append(fcf)

    if wacc <= terminal_growth:
        return {"dcf_value": None, "fcf_list": fcf_list}

    terminal_value = fcf_list[-1] * (1 + terminal_growth) / (wacc - terminal_growth)

    dcf = 0.0
    for i, fcf in enumerate(fcf_list, start=1):
        dcf += fcf / ((1 + wacc) ** i)
    dcf += terminal_value / ((1 + wacc) ** years)

    return {"dcf_value": dcf, "fcf_list": fcf_list}

# =========================
# Offer & Synergies
# =========================

def build_offer(target: dict, premium_pct: float) -> dict:
    price = target["price"]
    shares = target["shares_out"]
    equity_value_current = price * shares
    offer_price = price * (1 + premium_pct)
    equity_value_offer = offer_price * shares

    total_debt = target["total_debt"] or 0
    cash = target["cash"] or 0
    net_debt = total_debt - cash
    ev_offer = equity_value_offer + net_debt

    return {
        "current_price": price,
        "offer_price": offer_price,
        "premium_pct": premium_pct,
        "equity_value_current": equity_value_current,
        "equity_value_offer": equity_value_offer,
        "ev_offer": ev_offer,
        "premium_amount": equity_value_offer - equity_value_current,
    }


def build_synergies_runrate(target: dict,
                            cost_synergy_pct: float,
                            rev_synergy_pct: float,
                            synergy_margin: float,
                            tax_rate: float) -> dict:
    revenue = target["revenue"] or 0
    annual_cost_synergies = cost_synergy_pct * revenue
    annual_rev_synergies = rev_synergy_pct * revenue
    annual_rev_synergy_profit = annual_rev_synergies * synergy_margin
    annual_synergy_ebit = annual_cost_synergies + annual_rev_synergy_profit
    annual_synergy_after_tax = annual_synergy_ebit * (1 - tax_rate)
    return {
        "annual_cost_synergies": annual_cost_synergies,
        "annual_rev_synergies": annual_rev_synergies,
        "annual_synergy_ebit": annual_synergy_ebit,
        "annual_synergy_after_tax": annual_synergy_after_tax,
    }


def build_synergy_npv(annual_synergy_after_tax: float,
                      wacc: float,
                      horizon_years: int = 10) -> dict:
    if annual_synergy_after_tax is None or annual_synergy_after_tax <= 0 or wacc <= 0:
        return {"synergy_npv": None, "ramp_profile": None}

    ramp_profile = [0.25, 0.6, 1.0]
    while len(ramp_profile) < horizon_years:
        ramp_profile.append(1.0)

    cash_flows = []
    for year in range(1, horizon_years + 1):
        factor = ramp_profile[year - 1]
        cf = annual_synergy_after_tax * factor
        cash_flows.append(cf)

    npv = 0.0
    for i, cf in enumerate(cash_flows, start=1):
        npv += cf / ((1 + wacc) ** i)

    return {"synergy_npv": npv, "ramp_profile": cash_flows}

# =========================
# Pro Forma & Accretion
# =========================

def build_pro_forma(buyer: dict,
                    target: dict,
                    offer: dict,
                    synergies_runrate: dict,
                    tax_rate: float,
                    debt_financing_pct: float,
                    stock_financing_pct: float) -> dict:
    deal_value = offer["equity_value_offer"]
    new_debt = deal_value * debt_financing_pct
    equity_issued_value = deal_value * stock_financing_pct
    cash_used = deal_value - new_debt - equity_issued_value

    buyer_net_debt0 = buyer["net_debt"] or 0
    target_net_debt0 = target["net_debt"] or 0

    pf_net_debt = buyer_net_debt0 + target_net_debt0 + new_debt

    pf_ebitda = (buyer["ebitda"] or 0) + (target["ebitda"] or 0) + synergies_runrate["annual_synergy_ebit"]

    cost_of_debt = estimate_cost_of_debt(buyer)
    annual_interest_new_debt = new_debt * cost_of_debt
    after_tax_interest_cost = annual_interest_new_debt * (1 - tax_rate)

    buyer_ni = buyer["net_income"] or 0
    target_ni = target["net_income"] or 0
    synergy_after_tax = synergies_runrate["annual_synergy_after_tax"]

    pf_net_income = buyer_ni + target_ni + synergy_after_tax - after_tax_interest_cost

    buyer_shares = buyer["shares_out"] or 0
    if buyer_shares <= 0:
        raise ValueError("Buyer shares outstanding missing or invalid.")

    new_shares_issued = 0.0
    if stock_financing_pct > 0:
        buyer_price = buyer["price"]
        if buyer_price and buyer_price > 0:
            new_shares_issued = equity_issued_value / buyer_price

    pf_shares = buyer_shares + new_shares_issued

    buyer_eps = buyer_ni / buyer_shares if buyer_shares > 0 else None
    pf_eps = pf_net_income / pf_shares if pf_shares > 0 else None

    accretion_pct = None
    if buyer_eps and buyer_eps != 0:
        accretion_pct = (pf_eps - buyer_eps) / buyer_eps

    pf_leverage = None
    if pf_ebitda and pf_ebitda > 0:
        pf_leverage = pf_net_debt / pf_ebitda

    return {
        "deal_value": deal_value,
        "new_debt": new_debt,
        "cash_used": cash_used,
        "equity_issued_value": equity_issued_value,
        "new_shares_issued": new_shares_issued,
        "pf_net_debt": pf_net_debt,
        "pf_ebitda": pf_ebitda,
        "pf_leverage": pf_leverage,
        "buyer_eps": buyer_eps,
        "pf_eps": pf_eps,
        "accretion_pct": accretion_pct,
        "after_tax_interest_cost": after_tax_interest_cost,
        "synergy_after_tax": synergy_after_tax,
    }

# =========================
# Strategic & Regulatory
# =========================

def strategic_fit_score(buyer: dict, target: dict) -> dict:
    score = 0
    sector_match = buyer["sector"] == target["sector"]
    industry_match = buyer["industry"] == target["industry"]

    if sector_match:
        score += 4
    if industry_match:
        score += 4

    return {
        "strategic_score": score,
        "sector_match": sector_match,
        "industry_match": industry_match,
    }


def regulatory_risk_score(buyer: dict, target: dict) -> dict:
    score = 0
    high_risk = False

    if buyer["market_cap"] and buyer["market_cap"] > 200_000_000_000:
        score += 3
    if buyer["sector"] == target["sector"]:
        score += 2

    if score >= 4:
        high_risk = True

    return {
        "reg_score": score,
        "high_risk": high_risk,
    }

# =========================
# Verdict
# =========================

def build_deal_verdict(offer: dict,
                       synergies_runrate: dict,
                       synergy_npv: dict,
                       pf: dict,
                       dcf: dict,
                       strat: dict,
                       reg: dict) -> dict:
    acc = pf["accretion_pct"]
    lev = pf["pf_leverage"]
    premium = offer["premium_pct"]
    premium_amount = offer["premium_amount"]
    synergy_after_tax = synergies_runrate["annual_synergy_after_tax"]
    synergy_npv_val = synergy_npv["synergy_npv"]
    dcf_value = dcf["dcf_value"]

    synergy_coverage_annual = None
    synergy_coverage_npv = None
    if premium_amount and premium_amount > 0:
        synergy_coverage_annual = synergy_after_tax / premium_amount
        if synergy_npv_val:
            synergy_coverage_npv = synergy_npv_val / premium_amount

    dcf_vs_offer = None
    if dcf_value:
        dcf_vs_offer = dcf_value - offer["equity_value_offer"]

    if acc is None or lev is None:
        verdict = "Inconclusive – missing key data ⚠️"
    else:
        if acc > 0.05 and (lev is None or lev <= 3.5) and premium <= 0.35:
            verdict = "Attractive deal candidate ✅"
        elif acc > 0 and (lev is None or lev <= 4.5) and premium <= 0.45:
            verdict = "Situational / Mixed – depends on strategic rationale ⚠️"
        else:
            verdict = "Financially challenging / unattractive deal ❌"

    return {
        "accretion_pct": acc,
        "pf_leverage": lev,
        "premium_pct": premium,
        "synergy_coverage_annual": synergy_coverage_annual,
        "synergy_coverage_npv": synergy_coverage_npv,
        "synergy_npv": synergy_npv_val,
        "dcf_value": dcf_value,
        "dcf_vs_offer": dcf_vs_offer,
        "strategic_score": strat["strategic_score"],
        "reg_score": reg["reg_score"],
        "reg_high_risk": reg["high_risk"],
        "verdict": verdict,
    }

# =========================
# Scenario Engine
# =========================

def run_case(buyer, target,
             premium_pct,
             debt_pct,
             stock_pct,
             tax_rate,
             cost_synergy_pct,
             rev_synergy_pct,
             synergy_margin):
    offer = build_offer(target, premium_pct)
    synergies_runrate = build_synergies_runrate(target, cost_synergy_pct, rev_synergy_pct, synergy_margin, tax_rate)
    wacc_target = estimate_wacc(target, tax_rate)
    synergy_npv = build_synergy_npv(synergies_runrate["annual_synergy_after_tax"], wacc_target)
    dcf = basic_dcf_target(target, tax_rate, wacc_target)
    pf = build_pro_forma(buyer, target, offer, synergies_runrate, tax_rate, debt_pct, stock_pct)
    strat = strategic_fit_score(buyer, target)
    reg = regulatory_risk_score(buyer, target)
    verdict = build_deal_verdict(offer, synergies_runrate, synergy_npv, pf, dcf, strat, reg)
    return {
        "offer": offer,
        "synergies": synergies_runrate,
        "synergy_npv": synergy_npv,
        "pf": pf,
        "dcf": dcf,
        "verdict": verdict,
    }


def build_scenarios(buyer, target,
                    base_premium,
                    debt_pct,
                    stock_pct,
                    tax_rate,
                    base_cost_synergy_pct,
                    base_rev_synergy_pct,
                    base_synergy_margin):
    base = run_case(
        buyer, target,
        base_premium,
        debt_pct,
        stock_pct,
        tax_rate,
        base_cost_synergy_pct,
        base_rev_synergy_pct,
        base_synergy_margin,
    )

    bull = run_case(
        buyer, target,
        base_premium,
        debt_pct,
        stock_pct,
        tax_rate,
        base_cost_synergy_pct * 1.5,
        base_rev_synergy_pct * 1.5,
        min(base_synergy_margin + 0.10, 0.9),
    )

    bear = run_case(
        buyer, target,
        base_premium,
        debt_pct,
        stock_pct,
        tax_rate,
        base_cost_synergy_pct * 0.6,
        base_rev_synergy_pct * 0.6,
        max(base_synergy_margin - 0.10, 0.0),
    )

    return {"bull": bull, "base": base, "bear": bear}


def build_premium_sensitivity(buyer, target,
                              debt_pct,
                              stock_pct,
                              tax_rate,
                              cost_synergy_pct,
                              rev_synergy_pct,
                              synergy_margin):
    premiums = [0.10, 0.20, 0.30, 0.40, 0.50]
    rows = []
    for p in premiums:
        case = run_case(
            buyer, target,
            p,
            debt_pct,
            stock_pct,
            tax_rate,
            cost_synergy_pct,
            rev_synergy_pct,
            synergy_margin,
        )
        v = case["verdict"]
        rows.append({
            "premium": p,
            "accretion": v["accretion_pct"],
            "pf_leverage": v["pf_leverage"],
            "synergy_cov_npv": v["synergy_coverage_npv"],
        })
    return rows


def build_financing_heatmap(buyer, target,
                            tax_rate,
                            cost_synergy_pct,
                            rev_synergy_pct,
                            synergy_margin):
    debt_levels = [0.0, 0.25, 0.5, 0.75, 1.0]
    premiums = [0.10, 0.20, 0.30, 0.40, 0.50]
    heatmap = []
    for p in premiums:
        row = {"premium": p, "cells": []}
        for d in debt_levels:
            s = 1.0 - d
            case = run_case(
                buyer, target,
                p,
                d,
                s,
                tax_rate,
                cost_synergy_pct,
                rev_synergy_pct,
                synergy_margin,
            )
            acc = case["verdict"]["accretion_pct"]
            row["cells"].append(acc)
        heatmap.append(row)
    return {"debt_levels": debt_levels, "premiums": premiums, "rows": heatmap}

# =========================
# Excel Formatting
# =========================

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
bold_font = Font(bold=True)


def write_table(ws, start_row, start_col, headers, rows, formats=None, title=None):
    r = start_row
    c = start_col
    if title:
        cell = ws.cell(row=r, column=c, value=title)
        cell.font = Font(bold=True, size=12)
        r += 1

    for j, h in enumerate(headers):
        cell = ws.cell(row=r, column=c + j, value=h)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    r += 1

    for row in rows:
        for j, val in enumerate(row):
            cell = ws.cell(row=r, column=c + j, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if formats and j < len(formats) and formats[j]:
                cell.number_format = formats[j]
        r += 1

    return r


def autosize_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = cell.value
                if val is None:
                    continue
                val = str(val)
                if len(val) > max_length:
                    max_length = len(val)
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

# =========================
# Excel Workbook Builder
# =========================

def build_excel_workbook(
    buyer, target,
    base_case,
    scenarios,
    premium_sens,
    heatmap,
    filename="deal_screen.xlsx",
):
    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "Summary"

    offer = base_case["offer"]
    verdict = base_case["verdict"]
    synergies = base_case["synergies"]
    synergy_npv = base_case["synergy_npv"]
    pf = base_case["pf"]
    dcf = base_case["dcf"]

    buyer_rows = [
        ["Ticker", buyer["ticker"]],
        ["Name", buyer["name"]],
        ["Sector", buyer["sector"]],
        ["Industry", buyer["industry"]],
        ["Market Cap", buyer["market_cap"]],
        ["Enterprise Value", buyer["enterprise_value"]],
        ["Net Debt", buyer["net_debt"]],
        ["EBITDA (TTM)", buyer["ebitda"]],
    ]
    write_table(
        ws_sum, 2, 2,
        ["Buyer", "Value"],
        buyer_rows,
        formats=[None, "#,##0"],
        title="Buyer Overview",
    )

    target_rows = [
        ["Ticker", target["ticker"]],
        ["Name", target["name"]],
        ["Sector", target["sector"]],
        ["Industry", target["industry"]],
        ["Market Cap", target["market_cap"]],
        ["Enterprise Value", target["enterprise_value"]],
        ["Net Debt", target["net_debt"]],
        ["EBITDA (TTM)", target["ebitda"]],
        ["EV / Revenue", target["ev_rev"]],
        ["EV / EBITDA", target["ev_ebitda"]],
    ]
    write_table(
        ws_sum, 2, 5,
        ["Target", "Value"],
        target_rows,
        formats=[None, "#,##0.00"],
        title="Target Overview",
    )

    offer_rows = [
        ["Current Price", offer["current_price"]],
        ["Offer Price", offer["offer_price"]],
        ["Premium", offer["premium_pct"]],
        ["Equity Value (Now)", offer["equity_value_current"]],
        ["Equity Value (Offer)", offer["equity_value_offer"]],
        ["Enterprise Value @ Offer", offer["ev_offer"]],
    ]
    write_table(
        ws_sum, 14, 2,
        ["Offer Metric", "Value"],
        offer_rows,
        formats=[None, "#,##0.00"],
        title="Offer Summary",
    )

    sy_rows = [
        ["Cost Synergies (annual)", synergies["annual_cost_synergies"]],
        ["Revenue Synergies (annual)", synergies["annual_rev_synergies"]],
        ["Synergy EBIT (annual)", synergies["annual_synergy_ebit"]],
        ["Synergy After Tax (annual)", synergies["annual_synergy_after_tax"]],
        ["Synergy NPV", synergy_npv["synergy_npv"]],
    ]
    write_table(
        ws_sum, 14, 5,
        ["Synergy Metric", "Value"],
        sy_rows,
        formats=[None, "#,##0"],
        title="Synergies",
    )

    pf_rows = [
        ["New Debt Raised", pf["new_debt"]],
        ["Equity Issued Value", pf["equity_issued_value"]],
        ["New Shares Issued", pf["new_shares_issued"]],
        ["Pro-Forma Net Debt", pf["pf_net_debt"]],
        ["Pro-Forma EBITDA", pf["pf_ebitda"]],
        ["Pro-Forma Leverage", pf["pf_leverage"]],
        ["Buyer EPS (Standalone)", pf["buyer_eps"]],
        ["Pro-Forma EPS", pf["pf_eps"]],
        ["EPS Accretion", verdict["accretion_pct"]],
        ["Synergy Coverage (NPV / premium)", verdict["synergy_coverage_npv"]],
        ["DCF Equity Value (approx)", dcf["dcf_value"]],
        ["DCF minus Offer Equity Value", verdict["dcf_vs_offer"]],
        ["Strategic Fit Score (0–8)", verdict["strategic_score"]],
        ["Regulatory Risk Score", verdict["reg_score"]],
        ["Reg High Risk", verdict["reg_high_risk"]],
        ["Verdict", verdict["verdict"]],
    ]
    write_table(
        ws_sum, 26, 2,
        ["Metric", "Value"],
        pf_rows,
        formats=[None, "#,##0"],
        title="Pro-Forma & Verdict",
    )

    autosize_columns(ws_sum)

    ws_sc = wb.create_sheet("Scenarios")
    headers = ["Case", "Accretion", "PF Leverage", "Synergy NPV/Premium", "Verdict"]
    rows = []
    for label in ["bull", "base", "bear"]:
        case = scenarios[label]
        v = case["verdict"]
        rows.append([
            label.capitalize(),
            v["accretion_pct"],
            v["pf_leverage"],
            v["synergy_coverage_npv"],
            v["verdict"],
        ])
    write_table(
        ws_sc, 2, 2,
        headers,
        rows,
        formats=[None, "0.0%", "0.0", "0.0%", None],
        title="Scenario Analysis",
    )
    autosize_columns(ws_sc)

    ws_ps = wb.create_sheet("Premium_Sensitivity")
    headers = ["Premium", "Accretion", "PF Leverage", "Synergy NPV/Premium"]
    rows = []
    for row in premium_sens:
        rows.append([
            row["premium"],
            row["accretion"],
            row["pf_leverage"],
            row["synergy_cov_npv"],
        ])
    write_table(
        ws_ps, 2, 2,
        headers,
        rows,
        formats=["0.0%", "0.0%", "0.0", "0.0%"],
        title="Premium Sensitivity",
    )
    autosize_columns(ws_ps)

    ws_hm = wb.create_sheet("Financing_Heatmap")
    debt_levels = heatmap["debt_levels"]
    rows_hm = heatmap["rows"]

    header_row = ["Premium"] + [f"Debt {int(d*100)}%" for d in debt_levels]
    data_rows = []
    for r in rows_hm:
        row_vals = [r["premium"]]
        for acc in r["cells"]:
            row_vals.append(acc)
        data_rows.append(row_vals)

    write_table(
        ws_hm, 2, 2,
        header_row,
        data_rows,
        formats=["0.0%"] + ["0.0%"] * len(debt_levels),
        title="Accretion Heatmap",
    )
    autosize_columns(ws_hm)

    ws_raw = wb.create_sheet("Raw_Data")
    raw_rows = [
        ["Buyer Ticker", buyer["ticker"]],
        ["Target Ticker", target["ticker"]],
        ["Buyer Market Cap", buyer["market_cap"]],
        ["Buyer EV", buyer["enterprise_value"]],
        ["Target Market Cap", target["market_cap"]],
        ["Target EV", target["enterprise_value"]],
    ]
    write_table(
        ws_raw, 2, 2,
        ["Field", "Value"],
        raw_rows,
        formats=[None, "#,##0"],
        title="Inputs Snapshot",
    )
    autosize_columns(ws_raw)

    wb.save(filename)
    return filename

# =========================
# PyQt5 GUI
# =========================

class DealScreenApp(QWidget):
    def __init__(self):
        super().__init__()
        self.provider = YahooDataProvider()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Acquisition Screening Tool v5 - PyQt5")
        layout = QGridLayout()
        layout.setSpacing(8)

        row = 0

        layout.addWidget(QLabel("Buyer Ticker"), row, 0)
        self.buyer_edit = QLineEdit()
        layout.addWidget(self.buyer_edit, row, 1)
        row += 1

        layout.addWidget(QLabel("Target Ticker"), row, 0)
        self.target_edit = QLineEdit()
        layout.addWidget(self.target_edit, row, 1)
        row += 1

        layout.addWidget(QLabel("Offer Premium (%)"), row, 0)
        self.premium_edit = QLineEdit("30")
        layout.addWidget(self.premium_edit, row, 1)
        row += 1

        layout.addWidget(QLabel("New Debt Financing (%)"), row, 0)
        self.debt_edit = QLineEdit("60")
        layout.addWidget(self.debt_edit, row, 1)
        row += 1

        layout.addWidget(QLabel("Stock Financing (%)"), row, 0)
        self.stock_edit = QLineEdit("0")
        layout.addWidget(self.stock_edit, row, 1)
        row += 1

        layout.addWidget(QLabel("Tax Rate (%)"), row, 0)
        self.tax_edit = QLineEdit("25")
        layout.addWidget(self.tax_edit, row, 1)
        row += 1

        layout.addWidget(QLabel("Cost Synergies (% of target rev)"), row, 0)
        self.cost_syn_edit = QLineEdit("5")
        layout.addWidget(self.cost_syn_edit, row, 1)
        row += 1

        layout.addWidget(QLabel("Revenue Synergies (% of target rev)"), row, 0)
        self.rev_syn_edit = QLineEdit("3")
        layout.addWidget(self.rev_syn_edit, row, 1)
        row += 1

        layout.addWidget(QLabel("Synergy Margin (%)"), row, 0)
        self.margin_edit = QLineEdit("30")
        layout.addWidget(self.margin_edit, row, 1)
        row += 1

        self.run_button = QPushButton("Generate Excel")
        self.run_button.clicked.connect(self.on_generate_excel)
        layout.addWidget(self.run_button, row, 0, 1, 2)
        row += 1

        self.result_label = QLabel("")
        self.result_label.setWordWrap(True)
        layout.addWidget(self.result_label, row, 0, 1, 2)

        self.setLayout(layout)
        self.resize(480, 420)
        self.center_on_screen()

    def center_on_screen(self):
        frame = self.frameGeometry()
        screen = QApplication.desktop().screenNumber(QApplication.desktop().cursor().pos())
        center_point = QApplication.desktop().screenGeometry(screen).center()
        frame.moveCenter(center_point)
        self.move(frame.topLeft())

    def on_generate_excel(self):
        buyer_ticker = self.buyer_edit.text().strip().upper()
        target_ticker = self.target_edit.text().strip().upper()

        if not buyer_ticker or not target_ticker:
            QMessageBox.critical(self, "Error", "Buyer and Target tickers are required.")
            return

        premium_pct = pct_to_float(self.premium_edit.text(), 30.0)
        debt_pct = pct_to_float(self.debt_edit.text(), 60.0)
        stock_pct = pct_to_float(self.stock_edit.text(), 0.0)
        if debt_pct + stock_pct > 1.0:
            stock_pct = max(0.0, 1.0 - debt_pct)

        tax_rate = pct_to_float(self.tax_edit.text(), 25.0)
        cost_syn = pct_to_float(self.cost_syn_edit.text(), 5.0)
        rev_syn = pct_to_float(self.rev_syn_edit.text(), 3.0)
        margin = pct_to_float(self.margin_edit.text(), 30.0)

        try:
            buyer = self.provider.get_company_snapshot(buyer_ticker)
            target = self.provider.get_company_snapshot(target_ticker)

            base_case = run_case(
                buyer, target,
                premium_pct,
                debt_pct,
                stock_pct,
                tax_rate,
                cost_syn,
                rev_syn,
                margin,
            )

            scenarios = build_scenarios(
                buyer, target,
                premium_pct,
                debt_pct,
                stock_pct,
                tax_rate,
                cost_syn,
                rev_syn,
                margin,
            )

            premium_sens = build_premium_sensitivity(
                buyer, target,
                debt_pct,
                stock_pct,
                tax_rate,
                cost_syn,
                rev_syn,
                margin,
            )

            heatmap = build_financing_heatmap(
                buyer, target,
                tax_rate,
                cost_syn,
                rev_syn,
                margin,
            )

            default_name = f"deal_screen_{buyer_ticker}_{target_ticker}.xlsx"
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Save Excel Workbook As...",
                default_name,
                "Excel Files (*.xlsx)"
            )

            if not filename:
                return

            build_excel_workbook(
                buyer, target,
                base_case,
                scenarios,
                premium_sens,
                heatmap,
                filename=filename,
            )

            acc = base_case["verdict"]["accretion_pct"]
            lev = base_case["verdict"]["pf_leverage"]
            verdict_text = base_case["verdict"]["verdict"]

            self.result_label.setText(
                f"Excel generated: {os.path.basename(filename)}\n"
                f"Accretion: {fmt_pct(acc)} | PF Leverage: {lev:.2f}x\n"
                f"Verdict: {verdict_text}"
            )

            QMessageBox.information(self, "Success", "Excel file created successfully.")

        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = DealScreenApp()
    window.show()
    sys.exit(app.exec_())
